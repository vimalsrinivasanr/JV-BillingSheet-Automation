import pandas as pd
import os
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP


def d2(val):
    """Round val to exactly 2 decimal places using Decimal to eliminate float precision noise."""
    return float(Decimal(str(val)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

# ──────────────
#  CONFIG  — change these every month
# ──────────────
BILLING_FILE   = "Input Data.xlsx"          # billing sheet filename
MONTH_END_DATE = "28022026"                 # end-of-month date  DDMMYYYY
MONTH_LABEL    = "Feb'26"                   # label used in Document Header Text
# ──────────────

MAX_LINES_PER_JV = 999                      # SAP hard limit
COST_CENTER      = "682490C510"             # fixed for all rows
PROFIT_CENTER    = "682490C5"               # fixed for all rows
COMPANY_CODE     = 6000
DOC_TYPE         = "SA"
CURRENCY         = "INR"
CREDIT_ACCOUNT   = 500003
CREDIT_POSTING_KEY = 40
DEBIT_POSTING_KEY  = 50

GL_COLUMNS = {
    85: 742234,   # Payroll            (Excel col CH, 0-index 85 in openpyxl = col index 84 in pandas)
    86: 742238,   # Manager cost
    87: 742235,   # Leadership cost
    88: 742236,   # Desk cost
    89: 742237,   # Retirals
    90: 842028,   # Mark-up (TP)
}

# Billing sheet column positions (0-based pandas index after header row)
# Header is on row index 1 (Excel row 3), data starts row index 2
COL = {
    "workday_id":        0,    # A — Workday ID
    "emp_no":            1,    # B — GCC employee number
    "name":              2,    # C — Name
    "capability_center": 6,    # G — Capability Center
    "total_billable":   37,    # AL — Total billable amount
    "client":           38,    # AM — Client
    "cost_center_src":  40,    # AO — Cost center (source entity)
    "legal_entity":     41,    # AP — Legal Entity
    "classification":   42,    # AQ — Classification (Billable / Non Billable)
    "billed_status":    79,    # CB — Billed / Unbilled / Accruals
    "ic_code":          80,    # CC — IC Code  (= Ref Key 1)
    "customer_code":    81,    # CD — Customer Code
    "invoice_no":       82,    # CE — Invoice Number
    "emp_no_ref":       83,    # CF — EmpNo ref  (= Ref Key 3)
    "cap_center_ref":   84,    # CG — Capability Center ref  (= Ref Key 2)
    "gl_742234":        85,    # CH — Payroll amount
    "gl_742238":        86,    # CI — Manager amount
    "gl_742235":        87,    # CJ — Leadership amount
    "gl_742236":        88,    # CK — Desk cost amount
    "gl_742237":        89,    # CL — Retirals amount
    "gl_842028":        90,    # CM — Mark-up amount
}

GL_COL_NAMES = ["gl_742234", "gl_742238", "gl_742235", "gl_742236", "gl_742237", "gl_842028"]
GL_CODES     = [742234,       742238,       742235,       742236,       742237,       842028]


def load_billing_sheet(filepath):
    """Load billing sheet. Row 0 = GL category, Row 1 = real headers, Row 2+ = data."""
    print(f"  Loading billing sheet: {filepath}")
    raw = pd.read_excel(filepath, sheet_name="Billing sheet", header=None, dtype=str)

    # Row index 2 (Excel row 3) is the actual column-name row
    # Row index 3 onwards is data
    col_names = raw.iloc[2].tolist()
    data = raw.iloc[3:].copy()
    data.columns = col_names
    data = data.reset_index(drop=True)

    # Pull only the columns we need by position
    cols_needed = {name: data.columns[pos] for name, pos in COL.items()}
    df = pd.DataFrame()
    for field, col_label in cols_needed.items():
        df[field] = data.iloc[:, COL[field]]

    print(f"  Total rows loaded: {len(df)}")
    return df


def clean_and_filter(df):
    """Filter: AQ=Billable AND CB=Billed only. Sort by EmpNo (CF). Each JV must sum to zero."""
    # Strip whitespace from all columns
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()

    # 1. Drop completely empty rows (no Workday ID) — prevents trailing blank rows
    df = df[df["workday_id"].notna() & ~df["workday_id"].isin(["nan", "None", ""])]

    # 2. Apply the correct filter:
    #    AQ (classification) = "Billable"  AND  CB (billed_status) = "Billed"
    #    This is the only inclusion rule — Non Billable and Unbilled rows are excluded.
    is_billable = df["classification"] == "Billable"
    is_billed   = df["billed_status"] == "Billed"

    before = len(df)
    df = df[is_billable & is_billed].copy()
    print(f"  Rows after filter (AQ=Billable AND CB=Billed): {len(df)}  (excluded {before - len(df)} rows)")

    if len(df) == 0:
        print("  ERROR: No rows passed the filter. Check classification and billed_status columns.")
        import sys; sys.exit(1)

    # 3. Ensure invoice numbers are present
    df = df[df["invoice_no"].notna() & ~df["invoice_no"].isin(["", "nan", "None"])]
    print(f"  Rows with valid invoice numbers: {len(df)}")

    # 4. Use Workday ID as fallback for missing emp_no_ref (Ref Key 3 / CF column)
    df["emp_no_ref"] = df["emp_no_ref"].replace(["", "nan", "None", "NaN"], None)
    df["emp_no_ref"] = df["emp_no_ref"].fillna(df["workday_id"])

    # Convert GL amount columns to numeric
    for col in GL_COL_NAMES:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Sort by Invoice No. then EmpNo (CF column) to match reference file sequence
    print(f"  Sorting by Invoice No. and EmpNo (CF)...")
    df = df.sort_values(by=["invoice_no", "emp_no_ref"]).reset_index(drop=True)

    return df


def make_debit_row(serial, invoice, emp, gl_col, gl_code, doc_header):
    amount = emp[gl_col]
    # Removed: if amount == 0.0:
    # Removed:     return None
    return {
        "Reference":           serial,
        "Document Date":       MONTH_END_DATE,
        "Document Type":       DOC_TYPE,
        "Company Code":        COMPANY_CODE,
        "Posting Date":        MONTH_END_DATE,
        "Reference.1":         invoice,
        "Document Header Text": doc_header,
        "Currency":            CURRENCY,
        "Exchange rate":       None,
        "Amount":              d2(-abs(amount)),
        "Posting Key":         DEBIT_POSTING_KEY,
        "Account":             gl_code,
        "Special G/L ind.":    None,
        "Cost Center":         COST_CENTER,
        "Internal Order":      None,
        "Profit Center":       PROFIT_CENTER,
        "Business Area":       None,
        "Assignment Number (20)": invoice,
        "Item Text (50)":      doc_header,
        "Ref Key 1":           emp["ic_code"],
        "Ref Key 2":           emp["cap_center_ref"],
        "Ref Key 3 (20)":      emp["emp_no_ref"],
        "Material":            None, "Trading Partner": None,
        "Tax Code":            None, "Withholding tax code": None,
        "Withholding tax base amount in document currency": None,
        "Customer": None, "Contracts": None, "Revenue Period": None,
        "Core Consultant": None, "Revenue Month": None, "Reversal Date": None,
        "LEDGER": None, "WT CODE1": None, "WT Amount": None,
        "Inovice Receipt Date": MONTH_END_DATE,
    }


def make_credit_row(serial, invoice, ic_code, credit_amount, doc_header):
    return {
        "Reference":           serial,
        "Document Date":       MONTH_END_DATE,
        "Document Type":       DOC_TYPE,
        "Company Code":        COMPANY_CODE,
        "Posting Date":        MONTH_END_DATE,
        "Reference.1":         invoice,
        "Document Header Text": doc_header,
        "Currency":            CURRENCY,
        "Exchange rate":       None,
        "Amount":              d2(credit_amount),
        "Posting Key":         CREDIT_POSTING_KEY,
        "Account":             CREDIT_ACCOUNT,
        "Special G/L ind.":    None,
        "Cost Center":         COST_CENTER,
        "Internal Order":      None,
        "Profit Center":       PROFIT_CENTER,
        "Business Area":       None,
        "Assignment Number (20)": invoice,
        "Item Text (50)":      doc_header,
        "Ref Key 1":           ic_code,
        "Ref Key 2":           None, "Ref Key 3 (20)": None,
        "Material":            None, "Trading Partner": None,
        "Tax Code":            None, "Withholding tax code": None,
        "Withholding tax base amount in document currency": None,
        "Customer": None, "Contracts": None, "Revenue Period": None,
        "Core Consultant": None, "Revenue Month": None, "Reversal Date": None,
        "LEDGER": None, "WT CODE1": None, "WT Amount": None,
        "Inovice Receipt Date": MONTH_END_DATE,
    }


def build_jv_rows(df):
    """
    For each billed employee row:
      - Create 6 debit lines (one per GL code), skipping zeros
    For each unique invoice number:
      - Create 1 credit line (sum of all debit amounts for that invoice)

    Large invoices (>998 debit lines) are split across multiple JV serials,
    each with its own balancing credit line, so no JV entry ever exceeds 999 lines.
    Each unique invoice number (or part of it if split) will correspond to a new JV serial.
    """
    doc_header   = f"Revenue Reclass {MONTH_LABEL}"
    all_invoices = df["invoice_no"].unique()
    print(f"  Unique invoice numbers found: {len(all_invoices)}")

    rows = []
    serial_counter = 1 # This will be the current serial number to assign

    for invoice in all_invoices:
        inv_df = df[df["invoice_no"] == invoice]
        # Assuming ic_code is consistent for an invoice, pick the first one
        ic_code = inv_df.iloc[0]["ic_code"]

        # Build all debit rows for every employee in this invoice
        all_debits_for_invoice = []
        for _, emp in inv_df.iterrows():
            for gl_col, gl_code in zip(GL_COL_NAMES, GL_CODES):
                r2 = emp["cap_center_ref"]
                r3 = emp["emp_no_ref"]
                
                # We still need a temporary emp-like object or modified function to handle this
                row = {
                    "Reference":           0, # assigned later
                    "Document Date":       MONTH_END_DATE,
                    "Document Type":       DOC_TYPE,
                    "Company Code":        COMPANY_CODE,
                    "Posting Date":        MONTH_END_DATE,
                    "Reference.1":         invoice,
                    "Document Header Text": doc_header,
                    "Currency":            CURRENCY,
                    "Exchange rate":       None,
                    "Amount":              round(-abs(emp[gl_col]), 2),
                    "Posting Key":         DEBIT_POSTING_KEY,
                    "Account":             gl_code,
                    "Special G/L ind.":    None,
                    "Cost Center":         COST_CENTER,
                    "Internal Order":      None,
                    "Profit Center":       PROFIT_CENTER,
                    "Business Area":       None,
                    "Assignment Number (20)": invoice,
                    "Item Text (50)":      doc_header,
                    "Ref Key 1":           emp["ic_code"],
                    "Ref Key 2":           r2,
                    "Ref Key 3 (20)":      r3,
                    "Material":            None, "Trading Partner": None,
                    "Tax Code":            None, "Withholding tax code": None,
                    "Withholding tax base amount in document currency": None,
                    "Customer": None, "Contracts": None, "Revenue Period": None,
                    "Core Consultant": None, "Revenue Month": None, "Reversal Date": None,
                    "LEDGER": None, "WT CODE1": None, "WT Amount": None,
                    "Inovice Receipt Date": MONTH_END_DATE,
                }
                all_debits_for_invoice.append(row)

        if not all_debits_for_invoice:
            continue

        # ── Slice debits into batches that fit within the 999 limit ──
        # Each batch needs 1 credit line, so max debits per batch = 998
        MAX_DEBITS_PER_BATCH = MAX_LINES_PER_JV - 1

        i = 0
        while i < len(all_debits_for_invoice):
            # Assign the current serial_counter to this new JV entry
            current_serial = serial_counter

            # Take as many debit rows as fit into MAX_DEBITS_PER_BATCH
            batch_size = min(MAX_DEBITS_PER_BATCH, len(all_debits_for_invoice) - i)
            batch = all_debits_for_invoice[i : i + batch_size]

            # Update serial number on each debit row in this batch
            for r in batch:
                r["Reference"] = current_serial

            # Credit line for this batch — sum debits with Decimal for exact arithmetic
            batch_credit = sum(Decimal(str(abs(r["Amount"]))) for r in batch)
            credit = make_credit_row(current_serial, invoice, ic_code, float(batch_credit), doc_header)

            rows.append(credit)
            rows.extend(batch)

            # Add a blank row for readability between JV entries
            blank_row = {k: None for k in credit.keys()}
            rows.append(blank_row)

            # Increment serial_counter for the next JV entry (either for the same invoice or the next one)
            serial_counter += 1
            i += batch_size

    print(f"  Total output rows built: {len(rows)}")
    # Subtract 1 because serial_counter increments one last time after the last batch/invoice
    print(f"  JV entries (serial numbers) used: {serial_counter - 1}")
    return rows


def write_output(rows, out_path):
    """Write final JV rows to Excel in SAP upload column order, including the 2 extra header rows."""
    SAP_COLUMNS = [
        "Reference", "Document Date", "Document Type", "Company Code", "Posting Date",
        "Reference.1", "Document Header Text", "Currency", "Exchange rate", "Amount",
        "Posting Key", "Account", "Special G/L ind.", "Cost Center", "Internal Order",
        "Profit Center", "Business Area", "Assignment Number (20)", "Item Text (50)",
        "Ref Key 1", "Ref Key 2", "Ref Key 3 (20)", "Material", "Trading Partner",
        "Tax Code", "Withholding tax code", "Withholding tax base amount in document currency",
        "Customer", "Contracts", "Revenue Period", "Core Consultant", "Revenue Month",
        "Reversal Date", "LEDGER", "WT CODE1", "WT Amount", "Inovice Receipt Date",
    ]

    # Convert rows to DataFrame
    df_data = pd.DataFrame(rows)
    # The columns in 'rows' dictionaries match SAP_COLUMNS except Reference_Invoice which was renamed in the dictionary during build_jv_rows
    # Let's fix column names to match SAP_COLUMNS
    df_data.columns = SAP_COLUMNS

    # Row 1 Labels
    row1 = [
        "Unique S.No.", "End of Month date", "SA", COMPANY_CODE, "End of Month date",
        "Invoice No.", f"Revenue Reclass {MONTH_LABEL}", None, None, "Amount",
        "Based on the formula", "GL Codes", None, "Standard", None, "Standard", None,
        "Invoice No.", f"Revenue Reclass {MONTH_LABEL}", "IC Code", "Capablity Center", "EmpNo.",
        None, None, None, None, None, None, None, None, None, None, None, None, None, None, "End of Month date"
    ]
    # Row 2 is all empty (None)
    row2 = [None] * len(SAP_COLUMNS)

    # Convert to DataFrame for insertion
    df_h1 = pd.DataFrame([row1], columns=SAP_COLUMNS)
    df_h2 = pd.DataFrame([row2], columns=SAP_COLUMNS)
    df_h3 = pd.DataFrame([SAP_COLUMNS], columns=SAP_COLUMNS) # Row 3 is the header row

    # Concatenate: Row 1 + Row 2 + Header Row + Data
    # Note: df_data already has the data. We want Row 1, Row 2, Row 3 (Actual Column Names), then Data.
    # Actually, pd.to_excel with header=True will write Row 3.
    # So we just need to prepend Row 1 and Row 2.
    
    final_df = pd.concat([df_h1, df_h2, df_data], ignore_index=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # header=False because Row 3 (the labels) is already in the dataframe as the 3rd row or by to_excel
        # Wait, if we want the 3rd row to be the column names of the data below it:
        # Row 1: labels
        # Row 2: empty
        # Row 3: Reference, Document Date...
        # Row 4: Data...
        # The easiest way:
        df_h1.to_excel(writer, sheet_name="JV", index=False, header=False, startrow=0)
        df_h2.to_excel(writer, sheet_name="JV", index=False, header=False, startrow=1)
        df_data.to_excel(writer, sheet_name="JV", index=False, header=True, startrow=2)

    # POST-PROCESS: Use openpyxl to rewrite Amount column with clean 2dp floats.
    # This ensures Excel's SUM sees exact values with no hidden binary precision residual.
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["JV"]

    # Row 3 is the header row. Find the Amount column index.
    amt_col_idx = None
    for cell in ws[3]:
        if str(cell.value).strip() == "Amount":
            amt_col_idx = cell.column
            break

    if amt_col_idx:
        last_data_row = 3  # header is row 3, data starts row 4
        for row_cells in ws.iter_rows(min_row=4, min_col=amt_col_idx, max_col=amt_col_idx):
            cell = row_cells[0]
            if cell.value is not None and cell.value != '':
                try:
                    clean_val = float(Decimal(f'{float(cell.value):.2f}'))
                    cell.value = clean_val
                    cell.number_format = '#,##0.00'
                    last_data_row = cell.row
                except (ValueError, TypeError, Exception):
                    pass

        # Inject ROUND(SUM(...), 2) formula in the first empty row below data
        # This ensures Excel shows exactly 0.00 instead of the floating-point residual
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(amt_col_idx)
        formula_row = last_data_row + 1
        formula_cell = ws.cell(row=formula_row, column=amt_col_idx)
        formula_cell.value = f'=ROUND(SUM({col_letter}4:{col_letter}{last_data_row}),2)'
        formula_cell.number_format = '#,##0.00'
        print(f"  Injected balance check formula at {col_letter}{formula_row}: =ROUND(SUM({col_letter}4:{col_letter}{last_data_row}),2)")

    wb.save(out_path)
    wb.close()
    print(f"  Output saved: {out_path}")


def validate(rows):
    """Basic sanity checks before writing the file."""
    print("\n--- Validation ---")
    df = pd.DataFrame(rows)
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")

    # Vectorized check: sum all amounts per serial — should be 0
    errors = 0
    serial_sums = df.groupby("Reference")["Amount"].sum().round(2)
    bad = serial_sums[serial_sums.abs() > 0.05]
    if not bad.empty:
        for serial, diff in bad.items():
            print(f"  MISMATCH serial={serial} diff={diff}")
        errors += len(bad)

    # Check no JV serial exceeds 999 lines
    serial_counts = df.groupby("Reference").size()
    over_limit = serial_counts[serial_counts > MAX_LINES_PER_JV]
    if not over_limit.empty:
        for serial, count in over_limit.items():
            print(f"  999 VIOLATION serial={serial} has {count} lines")
        errors += len(over_limit)

    if errors == 0:
        print(f"  All checks passed. {len(df)} rows ready for SAP upload.")
    else:
        print(f"  {errors} issue(s) found — review before uploading.")

    # Summary per invoice
    print("\n--- Invoice summary ---")
    summary = df[df["Posting Key"] == CREDIT_POSTING_KEY][["Assignment Number (20)", "Ref Key 1", "Amount"]].copy()
    summary.columns = ["Invoice No.", "IC Code", "Total Credit"]
    print(summary.to_string(index=False))


def main():
    print("=" * 55)
    print("  GCC SAP JV Upload Automation")
    print(f"  Month: {MONTH_LABEL}   Date: {MONTH_END_DATE}")
    print("=" * 55)

    if not os.path.exists(BILLING_FILE):
        print(f"\nERROR: Cannot find '{BILLING_FILE}'. Make sure it is in the same folder as this script.")
        sys.exit(1)

    print("\n[1] Loading billing sheet...")
    df = load_billing_sheet(BILLING_FILE)

    print("\n[2] Filtering Billed rows only...")
    df = clean_and_filter(df)

    print("\n[3] Building JV rows...")
    rows = build_jv_rows(df)

    print("\n[4] Validating output...")
    validate(rows)

    print("\n[5] Writing output file...")
    out_filename = f"SAP_JV_Upload_{MONTH_LABEL.replace(chr(39),'')}.xlsx"
    write_output(rows, out_filename)

    print("\nDone!")
    print(f"  File ready: {out_filename}")
    print("  Open the file, spot-check 2-3 employees against the billing sheet, then upload to SAP.")


if __name__ == "__main__":
    main()