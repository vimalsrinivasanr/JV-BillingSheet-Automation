import pandas as pd
import os
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

class JVEngine:
    """Core logic engine to handle data cleaning, formula calculation, and JV generation."""

    def __init__(self, config=None):
        self.config = config or {}
        self.MONTH_LABEL = self.config.get("MONTH_LABEL", "Feb'26")
        self.MONTH_END_DATE = self.config.get("MONTH_END_DATE", "28022026")
        self.COMPANY_CODE = self.config.get("COMPANY_CODE", 6000)
        self.CURRENCY = self.config.get("CURRENCY", "INR")
        self.DOC_TYPE = self.config.get("DOC_TYPE", "SA")
        self.MAX_LINES_PER_JV = 999
        self.CREDIT_ACCOUNT = 500003
        self.DEBIT_POSTING_KEY = "50"
        self.CREDIT_POSTING_KEY = "40"
        self.COST_CENTER = self.config.get("COST_CENTER", "682490C510")
        self.PROFIT_CENTER = self.config.get("PROFIT_CENTER", "682490C5")

        self.GL_COL_NAMES = ["gl_742234", "gl_742238", "gl_742235", "gl_742236", "gl_742237", "gl_842028"]
        self.GL_CODES = [742234, 742238, 742235, 742236, 742237, 842028]
        self.last_filter_check = None

    def _load_normalized_data(self, filepath):
        """Load Stage-1 normalized file (sheet: Normalized) and map to engine schema."""
        norm = pd.read_excel(filepath, sheet_name="Normalized", dtype=str).fillna("")

        def col(name, fallback=""):
            if name in norm.columns:
                return norm[name]
            if isinstance(fallback, str):
                return pd.Series([fallback] * len(norm), index=norm.index)
            return fallback

        df = pd.DataFrame()
        df["workday_id"] = col("Workday ID")
        df["cap_center"] = col("Capability Center")
        df["legal_entity"] = col("Legal Entity")
        df["classification"] = col("Classification")
        df["billed_status"] = col("Billed/ Unbilled")
        df["ic_code"] = col("IC Code", "UNKNOWN")
        df["invoice_no"] = col("Invoice No.")
        df["emp_no_ref"] = col("EmpNo (ref)", df["workday_id"])
        df["cap_center_ref"] = col("Capability Center (ref)", df["cap_center"])

        # GL columns are expected from Stage 1 output
        df["gl_742234"] = col("Recharge - Payroll", "0")
        df["gl_742238"] = col("Recharge - Manager", "0")
        df["gl_742235"] = col("Recharge - Leadership", "0")
        df["gl_742236"] = col("Recharge - Desk Cost", "0")
        df["gl_742237"] = col("Recharge - Retirals", "0")
        df["gl_842028"] = col("Mark up", "0")

        return df

    def _load_legacy_billing_data(self, filepath, log_callback):
        """Legacy fallback loader for raw billing sheet with historical index assumptions."""
        raw = pd.read_excel(filepath, sheet_name="Billing sheet", header=None, dtype=str)
        col_names = raw.iloc[2].tolist()
        data = raw.iloc[3:].copy()
        data.columns = col_names
        data = data.reset_index(drop=True).fillna("")

        mapping = {
            "workday_id": 0, "cap_center": 6, "legal_entity": 41, "classification": 42,
            "billed_status": 79, "ic_code": 80, "invoice_no": 82, "emp_no_ref": 83, "cap_center_ref": 84
        }

        df = pd.DataFrame()
        df["workday_id"] = data.iloc[:, mapping["workday_id"]]
        df["cap_center"] = data.iloc[:, mapping["cap_center"]]
        df["legal_entity"] = data.iloc[:, mapping["legal_entity"]]
        df["classification"] = data.iloc[:, mapping["classification"]]
        df["billed_status"] = data.iloc[:, mapping["billed_status"]]
        df["ic_code"] = data.iloc[:, mapping["ic_code"]]
        df["invoice_no"] = data.iloc[:, mapping["invoice_no"]]
        df["emp_no_ref"] = data.iloc[:, mapping["emp_no_ref"]]
        df["cap_center_ref"] = data.iloc[:, mapping["cap_center_ref"]]

        if len(data.columns) > 85:
            for i, name in enumerate(self.GL_COL_NAMES):
                df[name] = pd.to_numeric(data.iloc[:, 85 + i], errors="coerce").fillna(0.0)
        else:
            log_callback("Data Check: Missing extended columns. Calculating from raw A-AQ...")
            df = self.calculate_virtual_columns(data, df)

        return df

    def d2(self, val):
        """Round val to exactly 2 decimal places using Decimal."""
        if val is None or pd.isna(val): return 0.0
        try:
            return float(Decimal(str(val)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        except:
            return 0.0

    def calculate_virtual_columns(self, raw_data, df):
        """Internal rules engine for missing columns CB-CM."""
        def s(idx): 
            if idx >= len(raw_data.columns): return pd.Series([0.0]*len(df))
            return pd.to_numeric(raw_data.iloc[:, idx], errors="coerce").fillna(0.0)

        df["billed_status"] = "Billed" 
        df["ic_code"] = raw_data.iloc[:, 80] if len(raw_data.columns) > 80 else "UNKNOWN"
        df["invoice_no"] = raw_data.iloc[:, 82] if len(raw_data.columns) > 82 else ""
        df["emp_no_ref"] = df["workday_id"]
        df["cap_center_ref"] = df["cap_center"]

        # Financial Formulas (Summing A-AQ raw data)
        df["gl_742234"] = s(14) + s(16) + s(19) + s(20) + s(21) + s(22) + s(23)
        df["gl_742238"] = s(26) + s(30)
        df["gl_742235"] = s(25)
        df["gl_742236"] = s(35)
        df["gl_742237"] = s(17) + s(18)
        df["gl_842028"] = s(34) + s(36)
        return df

    def run_processing(self, filepath, log_callback=print, api_key=None):
        """Full processing pipeline."""
        log_callback(f"Loading {os.path.basename(filepath)}...")
        xls = pd.ExcelFile(filepath)
        if "Normalized" in xls.sheet_names:
            log_callback("Stage 2: Using Stage-1 normalized sheet.")
            df = self._load_normalized_data(filepath)
        elif "Billing sheet" in xls.sheet_names:
            log_callback("Stage 2: Normalized sheet not found. Using legacy raw-sheet fallback.")
            df = self._load_legacy_billing_data(filepath, log_callback)
        else:
            raise ValueError("Input file must contain 'Normalized' or 'Billing sheet' sheet.")

        log_callback("Filtering and reconciling data...")
        pre = df.copy()
        for col in ["workday_id", "cap_center", "legal_entity", "classification", "billed_status", "ic_code", "invoice_no", "emp_no_ref", "cap_center_ref"]:
            df[col] = df[col].astype(str).str.strip()
        for col in ["workday_id", "classification", "billed_status", "invoice_no"]:
            pre[col] = pre[col].astype(str).str.strip()

        df = df[df["workday_id"].isin(["", "nan", "None"]) == False]
        df = df[
            (df["classification"].str.lower() == "billable") &
            (df["billed_status"].str.lower() == "billed")
        ]
        df = df[~df["invoice_no"].isin(["", "nan", "None"])]

        self.last_filter_check = self._build_filter_check(pre, df)
        log_callback(
            f"Filter Check: total={self.last_filter_check['total_rows']}, "
            f"used={self.last_filter_check['rows_used_for_jv']}, "
            f"excluded={self.last_filter_check['excluded_rows']}"
        )
        
        for col in self.GL_COL_NAMES:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            
        df = df.sort_values(by=["invoice_no", "workday_id"]).reset_index(drop=True)
        return self._build_rows(df)

    def _build_filter_check(self, pre_df, filtered_df):
        """Create summary metrics for filter validation."""
        total = len(pre_df)
        valid_workday = (~pre_df["workday_id"].isin(["", "nan", "None"]))
        billable = (pre_df["classification"].str.lower() == "billable")
        billed = (pre_df["billed_status"].str.lower() == "billed")
        invoice_present = (~pre_df["invoice_no"].isin(["", "nan", "None"]))

        used_mask = valid_workday & billable & billed & invoice_present
        excluded = pre_df[~used_mask].copy()

        return {
            "total_rows": int(total),
            "rows_with_workday": int(valid_workday.sum()),
            "billable_rows": int(billable.sum()),
            "billed_rows": int(billed.sum()),
            "rows_with_invoice": int(invoice_present.sum()),
            "rows_used_for_jv": int(used_mask.sum()),
            "excluded_rows": int((~used_mask).sum()),
            "excluded_preview": excluded[["workday_id", "classification", "billed_status", "invoice_no"]].head(25),
            "filtered_unique_invoices": int(filtered_df["invoice_no"].astype(str).nunique()) if len(filtered_df) else 0,
        }

    def _get_full_row(self):
        """Returns a dict with all 37 SAP columns initialized to None."""
        cols = [
            "Reference", "Document Date", "Document Type", "Company Code", "Posting Date",
            "Reference.1", "Document Header Text", "Currency", "Exchange rate", "Amount",
            "Posting Key", "Account", "Special G/L ind.", "Cost Center", "Internal Order",
            "Profit Center", "Business Area", "Assignment Number (20)", "Item Text (50)",
            "Ref Key 1", "Ref Key 2", "Ref Key 3 (20)", "Material", "Trading Partner",
            "Tax Code", "Withholding tax code", "Withholding tax base amount in document currency",
            "Customer", "Contracts", "Revenue Period", "Core Consultant", "Revenue Month",
            "Reversal Date", "LEDGER", "WT CODE1", "WT Amount", "Inovice Receipt Date",
        ]
        return {c: None for c in cols}

    def _build_rows(self, df):
        doc_header = f"Revenue Reclass {self.MONTH_LABEL}"
        all_invoices = df["invoice_no"].unique()
        rows = []
        serial_counter = 1

        for inv in all_invoices:
            inv_df = df[df["invoice_no"] == inv]
            ic_code = inv_df.iloc[0]["ic_code"]
            debits = []
            for _, emp in inv_df.iterrows():
                for gl_col, gl_code in zip(self.GL_COL_NAMES, self.GL_CODES):
                    if abs(emp[gl_col]) < 0.01: continue
                    r = self._get_full_row()
                    r.update({
                        "Reference": 0, "Document Date": self.MONTH_END_DATE, "Document Type": self.DOC_TYPE,
                        "Company Code": self.COMPANY_CODE, "Posting Date": self.MONTH_END_DATE,
                        "Reference.1": inv, "Document Header Text": doc_header, "Currency": self.CURRENCY,
                        "Amount": self.d2(-abs(emp[gl_col])), "Posting Key": self.DEBIT_POSTING_KEY,
                        "Account": gl_code, "Cost Center": self.COST_CENTER, "Profit Center": self.PROFIT_CENTER,
                        "Assignment Number (20)": inv, "Item Text (50)": doc_header,
                        "Ref Key 1": ic_code, "Ref Key 2": emp["cap_center_ref"], "Ref Key 3 (20)": emp["emp_no_ref"],
                        "Inovice Receipt Date": self.MONTH_END_DATE
                    })
                    debits.append(r)

            if not debits: continue
            
            # Batching logic
            batch_max = self.MAX_LINES_PER_JV - 1
            i = 0
            while i < len(debits):
                batch = debits[i:i+batch_max]
                batch_credit = sum(Decimal(str(abs(d["Amount"]))) for d in batch)
                
                cr = self._get_full_row()
                cr.update({
                    "Reference": int(serial_counter), "Document Date": self.MONTH_END_DATE, "Document Type": self.DOC_TYPE,
                    "Company Code": self.COMPANY_CODE, "Posting Date": self.MONTH_END_DATE,
                    "Reference.1": inv, "Document Header Text": doc_header, "Currency": self.CURRENCY,
                    "Amount": self.d2(float(batch_credit)), "Posting Key": self.CREDIT_POSTING_KEY,
                    "Account": self.CREDIT_ACCOUNT, "Cost Center": self.COST_CENTER, "Profit Center": self.PROFIT_CENTER,
                    "Assignment Number (20)": inv, "Item Text (50)": doc_header,
                    "Ref Key 1": ic_code, "Inovice Receipt Date": self.MONTH_END_DATE
                })
                for d in batch: d["Reference"] = int(serial_counter)
                
                rows.append(cr)
                rows.extend(batch)
                rows.append({k: None for k in cr.keys()}) # spacer
                serial_counter += 1
                i += batch_max
        return rows

    def write_excel(self, rows, out_path, log_callback=print):
        """Write final JV rows to Excel in SAP upload column order, including the 2 extra header rows."""
        SAP_COLUMNS = list(self._get_full_row().keys())

        # Create DataFrame from a list of rows that actually have data OR are intentional spacers
        df_data = pd.DataFrame(rows)
        
        if df_data.empty:
            raise ValueError("No data rows found to write. Check filters.")
            
        df_data.columns = SAP_COLUMNS

        # Row 1 Labels
        row1 = [
            "Unique S.No.", self.MONTH_END_DATE, self.DOC_TYPE, self.COMPANY_CODE, self.MONTH_END_DATE,
            "Invoice No.", f"Revenue Reclass {self.MONTH_LABEL}", None, None, "Amount",
            "Based on the formula", "GL Codes", None, "Standard", None, "Standard", None,
            "Invoice No.", f"Revenue Reclass {self.MONTH_LABEL}", "IC Code", "Capablity Center", "EmpNo.",
            None, None, None, None, None, None, None, None, None, None, None, None, None, None, self.MONTH_END_DATE
        ]
        df_h1 = pd.DataFrame([row1], columns=SAP_COLUMNS)
        df_h2 = pd.DataFrame([[None]*len(SAP_COLUMNS)], columns=SAP_COLUMNS)

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_h1.to_excel(writer, sheet_name="JV", index=False, header=False, startrow=0)
            df_h2.to_excel(writer, sheet_name="JV", index=False, header=False, startrow=1)
            df_data.to_excel(writer, sheet_name="JV", index=False, header=True, startrow=2)

        # POST-PROCESS: Openpyxl for precision and formula
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb["JV"]

        amt_col_idx = 10 # Column 'J'
        last_data_row = 3
        for row in range(4, ws.max_row + 1):
            cell = ws.cell(row=row, column=amt_col_idx)
            val = cell.value
            if val is not None and val != "":
                try: cell.value = float(Decimal(str(val)))
                except: pass
                last_data_row = row

        # Inject formula at the bottom
        check_row = ws.max_row + 1 if ws.max_row > 5 else 5
        ws.cell(row=check_row, column=amt_col_idx).value = f"=ROUND(SUM(J4:J{last_data_row}),2)"
        
        wb.save(out_path)

        if self.last_filter_check:
            wb = load_workbook(out_path)
            ws = wb.create_sheet("Filter Check")
            ws.append(["Metric", "Value"])
            ws.append(["Total source rows", self.last_filter_check["total_rows"]])
            ws.append(["Rows with Workday ID", self.last_filter_check["rows_with_workday"]])
            ws.append(["Rows with Classification = Billable", self.last_filter_check["billable_rows"]])
            ws.append(["Rows with Billed Status = Billed", self.last_filter_check["billed_rows"]])
            ws.append(["Rows with Invoice No.", self.last_filter_check["rows_with_invoice"]])
            ws.append(["Rows used for JV", self.last_filter_check["rows_used_for_jv"]])
            ws.append(["Rows excluded", self.last_filter_check["excluded_rows"]])
            ws.append(["Unique invoices in JV", self.last_filter_check["filtered_unique_invoices"]])
            ws.append([])
            ws.append(["Excluded Rows Preview", "", "", ""])
            ws.append(["workday_id", "classification", "billed_status", "invoice_no"])
            for row in self.last_filter_check["excluded_preview"].itertuples(index=False):
                ws.append(list(row))
            wb.save(out_path)

        log_callback(f"Balance check formula injected at J{check_row}")
        if self.last_filter_check:
            log_callback("Filter Check sheet created.")
