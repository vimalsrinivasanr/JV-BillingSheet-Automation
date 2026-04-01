"""
normalizer.py  –  Stage 1: Data Normalizer
===========================================
Reads ANY monthly billing sheet (inconsistent layout / renamed columns),
maps it to the February "Golden Template" column structure using both
header names AND actual data-value fingerprinting, calculates the
formula-derived GL columns, and writes a colour-coded Excel file for
human review.

Output file:  <input_name>_NORMALIZED.xlsx
  Sheet 1  "Normalized"       – clean data in February column order
  Sheet 2  "Mapping Report"   – how every column was discovered
  Sheet 3  "Warnings"         – missing / empty columns for user action
"""

import pandas as pd
import numpy as np
import os
import re
import warnings
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
#  GOLDEN SCHEMA
#  Each tuple: (canonical_name, feb_col_idx, data_type, [aliases])
#  Aliases are checked case-insensitively against the source file headers.
# ─────────────────────────────────────────────────────────────────────────────
GOLDEN_SCHEMA = [
    # ── Identity
    ("Workday ID",                  0,  "str",  ["workday id"]),
    ("EmpNo",                       1,  "str",  ["empno", "oracle id", "emp no"]),
    ("Name",                        2,  "str",  ["name"]),
    ("Date of joining",             3,  "date", ["date of joining"]),
    ("Last working day",            4,  "date", ["last working day"]),
    ("Grade",                       5,  "str",  ["grade ", "oracle level", "workday grade"]),
    ("Capability Center",           6,  "str",  ["capability center"]),
    ("ELT",                         7,  "str",  ["elt"]),
    ("Location",                    8,  "str",  ["location", "core location"]),
    ("Type",                        9,  "str",  ["type", "billing type"]),
    ("Employee Status",             10, "str",  ["billing", "employee status"]),
    ("Category of Billing",         11, "str",  ["category of billing"]),
    ("Markup %",                    12, "num",  ["markup %", "tp markup %"]),
    ("Billing days",                13, "num",  ["billing days"]),
    # ── Financial
    ("Fixed CTC",                   14, "num",  ["fixed ctc ", "fixed ctc", "monthly fixed pay", "ctc"]),
    ("% of Bonus",                  15, "num",  ["% of bonus"]),
    ("Bonus 2026",                  16, "num",  ["bonus 2026"]),
    ("Gratuity",                    17, "num",  ["gratuity"]),
    ("Leave Encashment",            18, "num",  ["leave encashment", "leave enc"]),
    ("PF admin Monthly",            19, "num",  ["pf admin monthly"]),
    ("EDLI Monthly",                20, "num",  ["edli monthly"]),
    ("Commision/ Joining bonus",    21, "num",  ["commision/ joining bonus", "commision/ other bonus",
                                                  "commission/ joining bonus"]),
    ("SBP Accrual",                 22, "num",  ["sbp accrual", "sbp cost"]),
    ("Insurance",                   23, "num",  ["insurance"]),
    ("Total Payroll",               24, "num",  ["total payroll"]),
    ("Leadership cost",             25, "num",  ["leadership cost"]),
    ("Managers desk fee",           26, "num",  ["managers desk fee", "vsl desk fee"]),
    ("Payroll excl Manager cost",   27, "num",  ["payroll excluding manager cost",
                                                  "payroll excluding vsl cost"]),
    ("Shiv Shankar's Cost",         28, "num",  ["shiv shankar's cost"]),
    ("Anusha's Cost",               29, "num",  ["anusha' cost", "anusha's cost"]),
    ("Manager cost",                30, "num",  ["manager cost", "vsl's cost", "vsls cost"]),
    ("Manager",                     31, "str",  ["manager", "vsl name"]),
    ("Manager ID",                  32, "str",  ["manager id", "vsl id"]),
    ("Total Cost excl desk cost",   33, "num",  ["total cost other than desk cost"]),
    ("TP on total cost",            34, "num",  ["tp on total cost"]),
    ("Gross Desk Cost",             35, "num",  ["gross desk cost"]),
    ("TP on desk cost",             36, "num",  ["tp on desk cost"]),
    ("Total billable amount",       37, "num",  ["total billable amount"]),
    ("Client",                      38, "str",  ["client "]),
    ("Client for sumifs",           39, "str",  ["client for sumifs"]),
    ("Cost center",                 40, "str",  ["cost center ", "cost center"]),
    ("Legal Entity",                41, "str",  ["legal entity"]),
    # ── Classification / Status
    ("Classification",              42, "str",  ["classification", "billable/ non billable",
                                                  "billable/non billable"]),
    ("FTE count",                   43, "num",  ["fte count"]),
    ("Comments",                    44, "str",  ["comments"]),
    # ── Supporting
    ("Service Area",                48, "str",  ["service area"]),
    ("LOB",                         49, "str",  ["lob"]),
    ("BU",                          50, "str",  ["bu"]),
    # ── LOP section
    ("Working_days",                55, "num",  ["working_days"]),
    ("Days_Worked",                 56, "num",  ["days_worked"]),
    ("LOP -1 month",                66, "num",  ["lop -1 month"]),
    ("LOP -2 Month",                67, "num",  ["lop -2 month"]),
    # ── Billing Status + Invoice  (CB–CE)
    ("Billed/ Unbilled",            79, "str",  ["billed/ unbilled", "billing status",
                                                  "billed/unbilled"]),
    ("IC Code",                     80, "str",  ["ic code", "ic code "]),
    ("Customer Code",               81, "str",  ["customer code"]),
    ("Invoice No.",                 82, "str",  ["invoice no.", "invoice no/ jv no",
                                                  "invoice no/jv no"]),
    ("EmpNo (ref)",                 83, "str",  ["empno"]),
    ("Capability Center (ref)",     84, "str",  ["capability center"]),
    # ── GL Recharge columns (CB–CM) – calculated, NOT copied from source
    ("Recharge - Payroll",          85, "num",  ["recharge  general cost - payroll"]),
    ("Recharge - Manager",          86, "num",  ["recharge  general cost - manager"]),
    ("Recharge - Leadership",       87, "num",  ["recharge  general cost - leadership cost"]),
    ("Recharge - Desk Cost",        88, "num",  ["recharge  general cost - desk cost"]),
    ("Recharge - Retirals",         89, "num",  ["recharge  general cost - retirals"]),
    ("Mark up",                     90, "num",  ["mark up"]),
    ("Diff",                        92, "num",  ["diff"]),
]

# Columns that must be calculated – do not map from source even if name matches
CALCULATED_COLS = {
    "Recharge - Payroll", "Recharge - Manager", "Recharge - Leadership",
    "Recharge - Desk Cost", "Recharge - Retirals", "Mark up", "Diff"
}

# Data-value fingerprint patterns (regex → canonical column name)
DATA_FINGERPRINTS = [
    (r"^W\d{7}$",          "Workday ID"),       # W0008891
    (r"^GCC\d{4}$",        "EmpNo"),            # GCC5670
    (r"^(NL|US|UK|IN|CA|AU|FR|DE|JP|SG)_",  "IC Code"),   # NL_RSH, US_SAP_L
    (r"^RG\d{8}$",         "Invoice No."),      # RG30001390
    (r"^\d{8}$",           "Invoice No."),      # 30001312  (Feb format)
]


# ─────────────────────────────────────────────────────────────────────────────
class BillingNormalizer:
    """
    Converts any monthly billing sheet to the February Golden Template format.
    Exposes a single public method: normalize(input_path) → output_path
    """

    def __init__(self, log_callback=print):
        self.log = log_callback
        self._warnings = []
        self._report   = []   # list of dicts for Mapping Report sheet

    # ─────────────────────────────────────────────────────────────────────────
    def normalize(self, input_path, output_path=None):
        """
        Main entry point.
        Returns (output_path, report_list).
        """
        self._warnings = []
        self._report   = []

        fname = os.path.basename(input_path)
        self.log(f"[Normalizer] Reading: {fname}")

        raw = pd.read_excel(input_path, sheet_name="Billing sheet",
                            header=None, dtype=str)
        self.log(f"[Normalizer] Raw sheet: {len(raw)} rows × {len(raw.columns)} cols")

        # ── 1. Find header row
        hdr_row, data_start = self._detect_header_row(raw)
        self.log(f"[Normalizer] Header at row {hdr_row}, data starts row {data_start}")

        headers   = [str(h).strip() if pd.notna(h) else "" for h in raw.iloc[hdr_row]]
        data      = raw.iloc[data_start:].copy().reset_index(drop=True)
        data.columns = headers

        # ── 2. Build column map: canonical → source_col_name (or None)
        col_map = self._build_col_map(headers, data)

        # ── 3. Assemble normalized DataFrame
        norm = self._assemble(col_map, data)

        # ── 4. Synthesise missing critical columns
        norm = self._synthesise(norm)

        # ── 5. Calculate GL recharge columns
        norm = self._calculate_gl(norm)

        # ── 6. Drop rows with no Workday ID
        before = len(norm)
        norm["Workday ID"] = norm["Workday ID"].astype(str).str.strip()
        norm = norm[~norm["Workday ID"].isin(["", "nan", "None"])]
        self.log(f"[Normalizer] Dropped {before - len(norm)} empty rows; "
                 f"{len(norm)} rows remain.")

        # ── 7. Write Excel
        if output_path is None:
            base, _ = os.path.splitext(input_path)
            output_path = base + "_NORMALIZED.xlsx"

        self._write_excel(norm, output_path)

        # ── 8. Summary log
        found   = sum(1 for r in self._report if r["status"] == "FOUND")
        synth   = sum(1 for r in self._report if r["status"] == "SYNTHESIZED")
        missing = sum(1 for r in self._report if r["status"] == "MISSING")
        self.log(f"[Normalizer] Mapping: {found} found, {synth} synthesised, {missing} missing")
        for r in self._report:
            if r["status"] == "MISSING":
                self.log(f"  ⚠  MISSING column: [{r['canonical']}]")
        for w in self._warnings:
            self.log(f"  ⚠  WARNING: {w}")
        self.log(f"[Normalizer] Output: {os.path.basename(output_path)}")

        return output_path, self._report

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 1 – Detect header row
    # ─────────────────────────────────────────────────────────────────────────
    def _detect_header_row(self, raw):
        """
        Scan up to row 20 for a row that contains at least 2 of the anchor
        tokens: 'workday id', 'empno', 'name'.
        """
        anchors = {"workday id", "empno", "name"}
        for i in range(min(20, len(raw))):
            row_lower = {str(v).strip().lower() for v in raw.iloc[i] if pd.notna(v)}
            if len(anchors & row_lower) >= 2:
                return i, i + 1
        self._warnings.append("Header row not auto-detected; defaulting to row 0.")
        return 0, 1

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 2 – Build column map
    # ─────────────────────────────────────────────────────────────────────────
    def _build_col_map(self, headers, data):
        """
        For each GOLDEN_SCHEMA column return the matching source column name
        (string key into data.columns) or None.

        Strategy (in order):
          A. Exact match (case-insensitive, strip whitespace)
          B. Alias match
          C. Data fingerprint scan (sample first 50 non-empty values per col)
        """
        # Lower-cased header → original header
        header_lc = {h.lower().strip(): h for h in headers if h.strip()}

        # Pre-compute data fingerprint: source_col → canonical (first hit)
        fp_map = self._fingerprint_columns(data, headers)

        col_map = {}  # canonical → source col name or None

        for (canon, feb_idx, dtype, aliases) in GOLDEN_SCHEMA:
            if canon in CALCULATED_COLS:
                col_map[canon] = None
                continue

            found_src = None
            method    = None

            # A. Exact / alias match on headers
            targets = [canon] + aliases
            for t in targets:
                key = t.lower().strip()
                if key in header_lc:
                    found_src = header_lc[key]
                    method    = f"header alias '{t}'"
                    break

            # B. Data fingerprint
            if not found_src and canon in fp_map.values():
                for src_col, fp_canon in fp_map.items():
                    if fp_canon == canon:
                        found_src = src_col
                        method    = "data fingerprint"
                        break

            status = "FOUND" if found_src else "MISSING"
            self._report.append({
                "canonical":  canon,
                "feb_idx":    feb_idx,
                "status":     status,
                "source_col": found_src,
                "method":     method or "",
            })
            col_map[canon] = found_src

        return col_map

    def _fingerprint_columns(self, data, headers):
        """
        For each source column, sample up to 50 non-null values and match
        against DATA_FINGERPRINTS. Returns { source_col_name → canonical }.
        """
        result = {}
        for h in headers:
            if not h.strip() or h not in data.columns:
                continue
            sample = data[h].dropna().astype(str).str.strip()
            sample = sample[sample != ""].head(50)
            for val in sample:
                for pattern, canon in DATA_FINGERPRINTS:
                    if re.match(pattern, val):
                        if h not in result:   # first hit wins
                            result[h] = canon
                        break
        return result

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 3 – Assemble normalised DataFrame
    # ─────────────────────────────────────────────────────────────────────────
    def _assemble(self, col_map, data):
        """Build DataFrame in February column order."""
        frames = {}
        for (canon, feb_idx, dtype, aliases) in GOLDEN_SCHEMA:
            src = col_map.get(canon)
            if src and src in data.columns:
                col = data[src]
                if dtype == "num":
                    frames[canon] = pd.to_numeric(col, errors="coerce")
                elif dtype == "date":
                    frames[canon] = pd.to_datetime(col, errors="coerce")
                else:
                    frames[canon] = col.astype(str).str.strip().replace("nan", "")
            else:
                frames[canon] = pd.Series([""] * len(data), dtype="object")
        return pd.DataFrame(frames)

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 4 – Synthesise missing critical columns
    # ─────────────────────────────────────────────────────────────────────────
    def _synthesise(self, df):
        # ── Billed/ Unbilled: derive from Invoice No. when column is blank
        status_blank = df["Billed/ Unbilled"].replace("", pd.NA).isna().all()
        inv_present  = not df["Invoice No."].replace("", pd.NA).isna().all()

        if status_blank and inv_present:
            self.log("[Normalizer] Synthesising 'Billed/ Unbilled' from Invoice column…")

            def _derive(val):
                v = str(val).strip().lower()
                if not v or v in ("nan", "none", ""):
                    return ""
                if "not billed" in v:
                    return "Unbillled"       # matches Feb spelling
                if "non billable" in v:
                    return "Non Billable"
                if "accrual" in v:
                    return "Accruals"
                return "Billed"             # actual invoice ID → Billed

            df["Billed/ Unbilled"] = df["Invoice No."].apply(_derive)
            for r in self._report:
                if r["canonical"] == "Billed/ Unbilled":
                    r["status"] = "SYNTHESIZED"
                    r["method"] = "derived from Invoice No."

        # ── EmpNo (ref): copy from Workday ID if blank
        if df["EmpNo (ref)"].replace("", pd.NA).isna().all():
            df["EmpNo (ref)"] = df["Workday ID"]
            for r in self._report:
                if r["canonical"] == "EmpNo (ref)":
                    r["status"] = "SYNTHESIZED"
                    r["method"] = "copied from Workday ID"

        # ── Capability Center (ref): copy from Capability Center if blank
        if df["Capability Center (ref)"].replace("", pd.NA).isna().all():
            df["Capability Center (ref)"] = df["Capability Center"]
            for r in self._report:
                if r["canonical"] == "Capability Center (ref)":
                    r["status"] = "SYNTHESIZED"
                    r["method"] = "copied from Capability Center"

        # ── IC Code warning if still empty
        if df["IC Code"].replace("", pd.NA).isna().all():
            self._warnings.append(
                "'IC Code' could not be found or is empty. "
                "JV Ref Key 1 will be blank. Please fill before JV generation."
            )

        return df

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 5 – Calculate GL recharge columns
    # ─────────────────────────────────────────────────────────────────────────
    def _calculate_gl(self, df):
        """
        Compute the formula-derived columns (equivalent to Feb CB–CM).
        Uses named columns from the normalised DataFrame.
        """
        def n(col):
            if col in df.columns:
                return pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            return pd.Series([0.0] * len(df), index=df.index)

        df["Recharge - Payroll"]    = n("Fixed CTC") + n("Bonus 2026") + \
                                      n("PF admin Monthly") + n("EDLI Monthly")
        df["Recharge - Manager"]    = n("Managers desk fee") + n("Manager cost")
        df["Recharge - Leadership"] = n("Leadership cost")
        df["Recharge - Desk Cost"]  = n("Gross Desk Cost")
        df["Recharge - Retirals"]   = n("Gratuity") + n("Leave Encashment")
        df["Mark up"]               = n("TP on total cost") + n("TP on desk cost")

        # Diff = Total billable – sum of all recharge columns (should be 0)
        sum_recharge = (df["Recharge - Payroll"] + df["Recharge - Manager"] +
                        df["Recharge - Leadership"] + df["Recharge - Desk Cost"] +
                        df["Recharge - Retirals"] + df["Mark up"])
        df["Diff"] = (n("Total billable amount") - sum_recharge).round(2)

        # Update report
        for r in self._report:
            if r["canonical"] in CALCULATED_COLS:
                r["status"] = "CALCULATED"
                r["method"] = "GL formula"

        self.log("[Normalizer] GL recharge columns calculated.")
        return df

    # ─────────────────────────────────────────────────────────────────────────
    #  Step 6 – Write colour-coded Excel output
    # ─────────────────────────────────────────────────────────────────────────
    def _write_excel(self, df, output_path):
        FILL_FOUND  = PatternFill("solid", fgColor="C6EFCE")  # green
        FILL_SYNTH  = PatternFill("solid", fgColor="FFEB9C")  # orange
        FILL_CALC   = PatternFill("solid", fgColor="BDD7EE")  # blue
        FILL_MISS   = PatternFill("solid", fgColor="FFC7CE")  # red
        HDR_FONT    = Font(bold=True, size=9)
        CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Status lookup for each column
        status_map = {r["canonical"]: r["status"] for r in self._report}

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Normalized", index=False)

        wb = load_workbook(output_path)
        ws = wb["Normalized"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = HDR_FONT
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = 19

            st = status_map.get(col_name, "MISSING")
            if st == "FOUND":
                cell.fill = FILL_FOUND
            elif st == "SYNTHESIZED":
                cell.fill = FILL_SYNTH
            elif st == "CALCULATED":
                cell.fill = FILL_CALC
            else:
                cell.fill = FILL_MISS

        ws.freeze_panes = "B2"

        # ── Mapping Report sheet
        rpt = pd.DataFrame(self._report).rename(columns={
            "canonical":  "Golden Column (Feb Format)",
            "feb_idx":    "Feb Index",
            "status":     "Status",
            "source_col": "Source Column Name",
            "method":     "Match Method",
        })
        rpt.to_excel(wb, sheet_name="Mapping Report", index=False)
        _autofit(wb["Mapping Report"])

        # ── Warnings sheet
        ws_w = wb.create_sheet("Warnings & Actions")
        ws_w.append(["Type", "Detail"])
        ws_w.append(["INFO", f"Source rows: {len(df)}  |  Columns: {len(df.columns)}"])
        ws_w.append(["INFO", "Green header = mapped | Orange = synthesised | Blue = calculated | Red = missing"])
        ws_w.append(["", ""])
        for r in self._report:
            if r["status"] == "MISSING":
                ws_w.append(["⚠ MISSING", f"'{r['canonical']}' (Feb col {r['feb_idx']}) was not found. Column is blank."])
        for w in self._warnings:
            ws_w.append(["⚠ WARNING", w])
        ws_w.append(["", ""])
        ws_w.append(["ACTION", "Review the 'Normalized' sheet. Fill red (missing) columns. Then run Stage 2 (JV Generator)."])
        _autofit(ws_w)

        wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────────────────────────────────────
#  CLI entry-point for quick testing
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "Input Data March.xlsx"
    norm = BillingNormalizer()
    out, rpt = norm.normalize(path)
    print(f"\nDone → {out}")
